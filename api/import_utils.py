import csv
import io
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

from django.contrib.auth.models import Group
from django.contrib.auth import get_user_model
from django.db import transaction

from openpyxl import load_workbook

from suv_tashish_crm.models import Client, Courier


User = get_user_model()


def _norm_phone(phone: str) -> str:
    if phone is None:
        return ""
    s = str(phone).strip()
    # remove spaces, dashes
    s = s.replace(" ", "").replace("-", "")
    # excel may store as float
    if s.endswith(".0"):
        s = s[:-2]
    # add +998 if 9 digits starting with 9x etc (basic)
    if s.startswith("998") and not s.startswith("+998"):
        s = "+" + s
    if s.startswith("9") and len(s) == 9:
        s = "+998" + s
    return s


def _ensure_group(name: str) -> Group:
    g, _ = Group.objects.get_or_create(name=name)
    return g


def _read_csv_bytes(data: bytes) -> List[Dict[str, str]]:
    text = data.decode("utf-8-sig", errors="ignore")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows = []
    for r in reader:
        rows.append({(k or "").strip().lower(): (v or "").strip() for k, v in r.items()})
    return rows


def _read_xlsx_bytes(data: bytes) -> List[Dict[str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [(str(h).strip().lower() if h is not None else "") for h in header]
    out: List[Dict[str, str]] = []
    for row in rows_iter:
        d = {}
        for i, key in enumerate(header):
            if not key:
                continue
            val = row[i] if i < len(row) else None
            d[key] = "" if val is None else str(val).strip()
        out.append(d)
    return out


def read_table_upload(uploaded_file) -> Tuple[List[Dict[str, str]], str]:
    name = (getattr(uploaded_file, "name", "") or "").lower()
    data = uploaded_file.read()
    if name.endswith(".csv"):
        return _read_csv_bytes(data), "csv"
    if name.endswith(".xlsx") or name.endswith(".xlsm") or name.endswith(".xltx"):
        return _read_xlsx_bytes(data), "xlsx"
    raise ValueError("UNSUPPORTED_FILE_TYPE")


@dataclass
class ImportResult:
    total: int
    created: int
    updated: int
    skipped: int
    errors: List[Dict]


def import_clients(rows: List[Dict[str, str]], default_password: str, mode: str = "upsert") -> ImportResult:
    client_group = _ensure_group("client")
    errors: List[Dict] = []
    created = updated = skipped = 0

    with transaction.atomic():
        seen = set()
        for idx, r in enumerate(rows, start=2):  # 1 header
            full_name = (r.get("full_name") or r.get("name") or "").strip()
            phone = _norm_phone(r.get("phone") or r.get("tel") or r.get("mobile") or "")
            email = (r.get("email") or "").strip()
            address = (r.get("address") or r.get("addr") or "").strip()
            lat = r.get("lat") or r.get("location_lat") or ""
            lon = r.get("lon") or r.get("location_lon") or ""

            if not full_name or not phone:
                skipped += 1
                errors.append({"row": idx, "error": "MISSING_FULL_NAME_OR_PHONE"})
                continue
            if phone in seen:
                skipped += 1
                errors.append({"row": idx, "error": "DUPLICATE_PHONE_IN_FILE", "phone": phone})
                continue
            seen.add(phone)

            user = User.objects.filter(username=phone).first()
            client = Client.objects.filter(phone=phone).first()

            if user is None and mode == "upsert" and client and client.user:
                user = client.user

            if user is None:
                user = User(username=phone, email=email or "")
                user.set_password(default_password)
                user.save()
                user.groups.add(client_group)
                created += 1
            else:
                # ensure group
                user.groups.add(client_group)
                # update email if provided
                if email and user.email != email:
                    user.email = email
                    user.save(update_fields=["email"])
                updated += 1

            if client is None:
                client = Client(phone=phone, full_name=full_name, user=user)
            else:
                if mode == "create_only":
                    # already exists -> skip
                    skipped += 1
                    continue
                client.user = user
                client.full_name = full_name or client.full_name

            if address:
                client.note = (client.note or "") + (("\n" if client.note else "") + f"Address: {address}")
            # location
            try:
                if lat:
                    client.location_lat = float(str(lat).replace(",", "."))
                if lon:
                    client.location_lon = float(str(lon).replace(",", "."))
            except Exception:
                errors.append({"row": idx, "error": "BAD_LAT_LON", "lat": lat, "lon": lon})

            client.must_change_password = True
            client.save()
    return ImportResult(total=len(rows), created=created, updated=updated, skipped=skipped, errors=errors)


def import_couriers(rows: List[Dict[str, str]], default_password: str, mode: str = "upsert") -> ImportResult:
    courier_group = _ensure_group("courier")
    errors: List[Dict] = []
    created = updated = skipped = 0

    with transaction.atomic():
        seen = set()
        for idx, r in enumerate(rows, start=2):
            full_name = (r.get("full_name") or r.get("name") or "").strip()
            phone = _norm_phone(r.get("phone") or r.get("tel") or r.get("mobile") or "")
            email = (r.get("email") or "").strip()

            if not full_name or not phone:
                skipped += 1
                errors.append({"row": idx, "error": "MISSING_FULL_NAME_OR_PHONE"})
                continue
            if phone in seen:
                skipped += 1
                errors.append({"row": idx, "error": "DUPLICATE_PHONE_IN_FILE", "phone": phone})
                continue
            seen.add(phone)

            user = User.objects.filter(username=phone).first()
            courier = Courier.objects.filter(phone=phone).first()
            if user is None and mode == "upsert" and courier and courier.user:
                user = courier.user

            if user is None:
                user = User(username=phone, email=email or "")
                user.set_password(default_password)
                user.save()
                user.groups.add(courier_group)
                created += 1
            else:
                user.groups.add(courier_group)
                if email and user.email != email:
                    user.email = email
                    user.save(update_fields=["email"])
                updated += 1

            if courier is None:
                courier = Courier(phone=phone, full_name=full_name, user=user)
            else:
                if mode == "create_only":
                    skipped += 1
                    continue
                courier.user = user
                courier.full_name = full_name or courier.full_name

            courier.must_change_password = True
            courier.save()
    return ImportResult(total=len(rows), created=created, updated=updated, skipped=skipped, errors=errors)
