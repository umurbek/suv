from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.db.models import Count, Sum, Q
from suv_tashish_crm.models import Order, Client, Courier, Region, Admin
from django.utils import timezone
import random
from decimal import Decimal
import datetime
import json
import csv
import os
import re
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.conf import settings


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def admin_dashboard(request):
    # Auth guard temporarily disabled; re-enable when ready
    today_orders = Order.objects.filter(created_at__date=timezone.localdate()).count()
    active_couriers = Courier.objects.filter(is_active=True).count()
    total_clients = Client.objects.count()
    debtors = Client.objects.filter(debt__gt=0).count()

    top_debtors_qs = Client.objects.filter(debt__gt=0).order_by('-debt')[:5]
    # prefer CSV names for clients that have generic 'Mijoz ...' placeholders
    static_regions = []
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []

    def _normalize_to_998(raw_digits: str):
        if not raw_digits:
            return None
        d = ''.join(ch for ch in raw_digits if ch.isdigit())
        if len(d) < 9:
            return None
        last9 = d[-9:]
        std = '998' + last9
        if len(std) == 12:
            return std
        return None

    csv_phone_map = {}
    csv_phone_by_last6 = {}
    for r in static_regions:
        p = (r.get('phone') or '').strip()
        norm = _normalize_to_998(p)
        if norm:
            csv_phone_map[norm] = r.get('name')
        digits_only = ''.join(ch for ch in p if ch.isdigit())
        if len(digits_only) >= 6:
            csv_phone_by_last6[digits_only[-6:]] = r.get('name')

    top_debtors = []
    for c in top_debtors_qs:
        name = c.full_name or ''
        # prefer CSV mapping when name is generic
        use_name = name
        if (not name) or name.strip().lower().startswith('mijoz'):
            norm = _normalize_to_998(c.phone or '')
            if norm and norm in csv_phone_map:
                use_name = csv_phone_map.get(norm)
            else:
                # try last-6 fallback
                digits_only = ''.join(ch for ch in (c.phone or '') if ch.isdigit())
                if len(digits_only) >= 6 and digits_only[-6:] in csv_phone_by_last6:
                    use_name = csv_phone_by_last6.get(digits_only[-6:])

        top_debtors.append({'id': c.id, 'full_name': use_name, 'region': getattr(c.region, 'name', None), 'debt': getattr(c, 'debt', 0)})
    # Build simple stats for the chart: last 7 days and last 30 days
    weekly_labels = []
    weekly_counts = []
    monthly_labels = []
    monthly_counts = []
    try:
        today = timezone.localdate()
        # last 7 days
        for i in range(6, -1, -1):
            day = today - datetime.timedelta(days=i)
            weekly_labels.append(day.strftime('%a'))
            weekly_counts.append(Order.objects.filter(created_at__date=day).count())

        # last 30 days: labels are day numbers (e.g., '01', '02', ...)
        for i in range(29, -1, -1):
            day = today - datetime.timedelta(days=i)
            monthly_labels.append(day.strftime('%d %b'))
            monthly_counts.append(Order.objects.filter(created_at__date=day).count())
    except Exception:
        weekly_labels = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
        weekly_counts = [0,0,0,0,0,0,0]
        monthly_labels = [f"{i+1}" for i in range(30)]
        monthly_counts = [0] * 30

    # Top clients by number of orders (for dashboard sidebar)
    try:
        top_clients_qs = (
            Order.objects.values('client__id', 'client__full_name')
            .annotate(orders_count=Count('id'))
            .order_by('-orders_count')[:5]
        )
        top_clients = [{'id': t['client__id'], 'full_name': t.get('client__full_name') or '—', 'orders_count': t.get('orders_count', 0)} for t in top_clients_qs]
    except Exception:
        top_clients = []
    # Top couriers by delivered orders (dynamic)
    try:
        top_couriers_qs = (
            Order.objects.filter(status='done')
            .values('courier__id', 'courier__full_name')
            .annotate(delivered=Count('id'))
            .order_by('-delivered')[:5]
        )
        top_couriers = [{'id': t.get('courier__id'), 'full_name': t.get('courier__full_name') or '—', 'delivered': t.get('delivered', 0)} for t in top_couriers_qs]
    except Exception:
        top_couriers = []

    # Orders per region (dynamic)
    try:
        region_stats_qs = (
            Order.objects.values('client__region__name')
            .annotate(total_orders=Count('id'))
            .order_by('-total_orders')[:10]
        )
        region_stats = [{'region': t.get('client__region__name') or '—', 'total_orders': t.get('total_orders', 0)} for t in region_stats_qs]
    except Exception:
        region_stats = []

    # Pending orders count (orders awaiting delivery) and today's revenue
    try:
        pending_count = Order.objects.filter(status__in=['assigned', 'delivering']).count()
    except Exception:
        pending_count = 0

    try:
        today = timezone.localdate()
    except Exception:
        today = datetime.date.today()

    try:
        revenue_qs = Order.objects.filter(
            Q(delivered_at__date=today) | (Q(status='done') & Q(created_at__date=today))
        )
        agg = revenue_qs.aggregate(total=Sum('payment_amount'))
        today_total = agg.get('total') or 0
        try:
            # convert Decimal to int for display in UZS (fallback to 0)
            today_revenue = int(today_total)
        except Exception:
            try:
                today_revenue = int(float(today_total))
            except Exception:
                today_revenue = 0
    except Exception:
        today_revenue = 0

    # Recent orders for the small table on dashboard
    try:
        recent_qs = Order.objects.select_related('client').order_by('-created_at')[:8]
        recent_orders = []
        for o in recent_qs:
            amt = None
            try:
                amt = int(o.payment_amount) if getattr(o, 'payment_amount', None) is not None else None
            except Exception:
                amt = None
            if amt is None:
                try:
                    amt = int(o.debt_change) if getattr(o, 'debt_change', None) is not None else 0
                except Exception:
                    amt = 0
                recent_orders.append({
                    'id': o.id,
                    'client_name': o.client.full_name if o.client else None,
                    'client': o.client.full_name if o.client else None,
                    'status': o.status,
                    'total_display': f"{amt:,}" if isinstance(amt, int) else str(amt),
                })
    except Exception:
        recent_orders = []

    context = {
        'today_orders': today_orders,
        'active_couriers': active_couriers,
        'total_clients': total_clients,
        'debtors': debtors,
        'top_debtors': top_debtors,
        'top_clients': top_clients,
        'pending_count': pending_count,
        'today_revenue': today_revenue,
        'recent_orders': recent_orders,
        'top_couriers': top_couriers,
        'region_stats': region_stats,
        'weekly_labels_json': json.dumps(weekly_labels, ensure_ascii=False),
        'weekly_data_json': json.dumps(weekly_counts),
        'monthly_labels_json': json.dumps(monthly_labels, ensure_ascii=False),
        'monthly_data_json': json.dumps(monthly_counts),
    }

    return render(request, 'admin/admin_dashboard.html', context)


def read_csv_data():
    regions = []
    csv_path = os.path.join(BASE_DIR, 'Volidam.csv')
    with open(csv_path, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        next(reader)  # Skip header
        for row in reader:
            regions.append({
                'name': row[0],
                'bottle': row[1],
                'location': row[2],
                'phone': row[3] if len(row) > 3 else ''
            })
    return regions


def regions_view(request):
    """Show regions list. Uses DB if available, falls back to CSV data."""
    static_regions = []
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []

    regions = []
    locations = []

    try:
        qs = Region.objects.all()
        for r in qs:
            name = getattr(r, 'name', '')
            bottle = getattr(r, 'bottle', '') if hasattr(r, 'bottle') else ''
            loc = getattr(r, 'location', '') if hasattr(r, 'location') else ''
            phone = getattr(r, 'phone', '') if hasattr(r, 'phone') else ''
            regions.append({'name': name, 'bottle': bottle, 'location': loc, 'phone': phone})
            if loc and loc not in locations:
                locations.append(loc)

        # If DB has no regions but CSV does, show CSV rows instead
        if not regions and static_regions:
            regions = static_regions
            seen = set()
            locations = []
            for r in static_regions:
                loc = (r.get('location') or '').strip()
                if loc and loc not in seen:
                    seen.add(loc)
                    locations.append(loc)
    except Exception:
        regions = static_regions
        seen = set()
        locations = []
        for r in regions:
            loc = (r.get('location') or '').strip()
            if loc and loc not in seen:
                seen.add(loc)
                locations.append(loc)

    return render(request, 'admin/regions.html', {'regions': regions, 'locations': locations})


def add_region(request):
    if request.method == 'POST':
        # Auth guard temporarily disabled; re-enable when ready
        try:
            import json
            payload = json.loads(request.body.decode('utf-8'))
            name = (payload.get('name') or '').strip()
            bottle = (payload.get('bottle') or '').strip()
            location = (payload.get('location') or '').strip()
            phone = (payload.get('phone') or '').strip()
            if not name:
                return JsonResponse({'status': 'error', 'message': "Hudud nomi kerak."})

            # create or get Region in DB
            region_obj, created = Region.objects.get_or_create(name=name)

            # We keep CSV metadata for backward compatibility but do not write to CSV here.
            return JsonResponse({'status': 'success', 'region': {'name': region_obj.name, 'bottle': bottle, 'location': location, 'phone': phone, 'created': created}})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def update_region(request):
    if request.method == 'POST':
        # Auth guard temporarily disabled; re-enable when ready
        # Expect JSON payload with old name and new fields
        try:
            import json
            payload = json.loads(request.body.decode('utf-8'))
            orig_name = payload.get('name') or payload.get('orig_name')
            new_name = payload.get('newName') or payload.get('name_new') or payload.get('new_name')
            new_bottle = payload.get('bottle')
            new_location = payload.get('location') or payload.get('newLocation')
            new_phone = payload.get('phone') or payload.get('newPhone')
            # Server-side phone validation: if provided, must be +998 followed by 9 digits
            if new_phone:
                new_phone = new_phone.strip()
                if not re.match(r'^\+998\d{9}$', new_phone):
                    return JsonResponse({'status': 'error', 'message': "Telefon formati noto'g'ri. Iltimos '+998XXXXXXXXX' formatida kiriting."})

            if not orig_name:
                return JsonResponse({'status': 'error', 'message': 'Orig name required.'})

            try:
                region_obj = Region.objects.get(name=orig_name)
            except Region.DoesNotExist:
                return JsonResponse({'status': 'not_found'})

            if new_name:
                region_obj.name = new_name
                region_obj.save()

            # We intentionally do not store bottle/location/phone on Region model (legacy CSV kept separately)
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def delete_region(request):
    if request.method == 'POST':
        # Auth guard temporarily disabled; re-enable when ready
        try:
            import json
            payload = json.loads(request.body.decode('utf-8'))
            name = (payload.get('name') or '').strip()
            if not name:
                return JsonResponse({'status': 'error', 'message': "Hudud nomi kerak."})

            try:
                region_obj = Region.objects.get(name=name)
                region_obj.delete()
                return JsonResponse({'status': 'success'})
            except Region.DoesNotExist:
                return JsonResponse({'status': 'not_found'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})


def couriers_view(request):
    # Auth guard temporarily disabled; re-enable when ready
    # Provide dynamic list of couriers to the template
    try:
        qs = Courier.objects.select_related('region').all().order_by('full_name')
        couriers = []
        for c in qs:
            status_label = 'Faol' if getattr(c, 'is_active', False) else 'Nofaol'
            couriers.append({
                'id': c.id,
                'full_name': c.full_name,
                'phone': c.phone,
                'region': getattr(c.region, 'name', ''),
                'is_active': getattr(c, 'is_active', False),
                'status_label': status_label,
            })
    except Exception:
        couriers = []

    # Pop any flash messages set by other views (add/edit/delete)
    flash_message = request.session.pop('flash_message', None)
    flash_type = request.session.pop('flash_type', None)

    # Render the modern styled template with optional flash
    return render(request, 'admin/couriers.html', {'couriers': couriers, 'flash_message': flash_message, 'flash_type': flash_type})


def region_clients_api(request):
    """Return JSON list of clients for a given region name.

    Query param: ?name=<region name>
    """
    name = request.GET.get('name') or request.POST.get('name')
    if not name:
        return JsonResponse({'status': 'error', 'message': 'Region name required.'}, status=400)

    try:
        clients_qs = Client.objects.filter(region__name=name)
    except Exception:
        clients_qs = Client.objects.none()

    clients = []
    for c in clients_qs:
        try:
            lat = float(getattr(c, 'location_lat', 0) or 0)
            lon = float(getattr(c, 'location_lon', 0) or 0)
        except Exception:
            lat = 0.0
            lon = 0.0

        clients.append({
            'id': c.id,
            'full_name': c.full_name,
            'phone': c.phone,
            'lat': lat,
            'lon': lon,
            'note': c.note or '',
        })

    return JsonResponse({'status': 'ok', 'clients': clients})


def clients_positions_api(request):
    """Return JSON list of all clients with lat/lon for admin map display."""
    try:
        qs = Client.objects.all()
    except Exception:
        qs = Client.objects.none()
    items = []
    for c in qs:
        try:
            lat = float(getattr(c, 'location_lat', 0) or 0)
            lon = float(getattr(c, 'location_lon', 0) or 0)
        except Exception:
            lat = 0.0; lon = 0.0
        # skip clients without a meaningful location
        if lat == 0 and lon == 0:
            continue
        items.append({'id': c.id, 'full_name': c.full_name, 'phone': c.phone, 'lat': lat, 'lon': lon, 'region': getattr(c.region, 'name', None)})
    return JsonResponse({'status': 'ok', 'clients': items})


def clients_view(request):
    # Auth guard temporarily disabled; re-enable when ready
    # Provide dynamic client profiles to template
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []
    static_map = {r.get('name'): r.get('location') for r in static_regions}
    clients_qs = Client.objects.select_related('region').all()
    clients = []
    for c in clients_qs:
        address = ''
        if c.region:
            address = static_map.get(c.region.name) or getattr(c.region, 'name', '')
        last_order = Order.objects.filter(client=c).order_by('-created_at').first()
        last_order_date = ''
        if last_order and getattr(last_order, 'created_at', None):
            try:
                last_order_date = last_order.created_at.strftime('%Y-%m-%d %H:%M')
            except Exception:
                last_order_date = str(last_order.created_at)
        clients.append({
            'id': c.id,
            'full_name': c.full_name,
            'phone': c.phone,
            'address': address,
            'debt': getattr(c, 'debt', 0),
            'bottle_balance': getattr(c, 'bottle_balance', 0),
            'last_order_date': last_order_date,
        })
    return render(request, 'admin/clients.html', {'clients': clients})


def clients_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = Client.objects.select_related("region").order_by("-id")
    if q:
        qs = qs.filter(full_name__icontains=q) | qs.filter(phone__icontains=q)

    return render(request, "admin/clients.html", {
        "clients": qs[:500],
        "q": q,
    })


from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from .models import AdminProfile

User = get_user_model()


# =========================
# PROFIL SAHIFASI
# =========================
from .models import AdminProfile


def profile_view(request):
    # Do not force redirect to login here; render profile page and
    # show editable fields only for authenticated users.
    admin_data = None
    if request.user.is_authenticated:
        # Admin profili yaratish yoki olish (default qiymatlar bilan)
        admin_data, created = AdminProfile.objects.get_or_create(
            user=request.user,
            defaults={
                'full_name': request.user.username,
                'phone': '',
                'role': 'Administrator'
            }
        )

    # Barcha adminlar ro'yxatini olish (template uchun)
    all_admins = Admin.objects.all()

    return render(request, 'admin/profile.html', {
        'admin_profile': admin_data,
        'user': request.user,
        'all_admins': all_admins,
        'total_admins': all_admins.count(),
    })

# =========================
# PROFILNI YANGILASH API
# =========================
@login_required(login_url='/admin_panel/login/')
def api_update_profile(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Faqat POST so‘rov qabul qilinadi'}, status=405)

    admin_name = request.POST.get('admin_name', '').strip()
    username = request.POST.get('username', '').strip()
    email = request.POST.get('email', '').strip()
    password = request.POST.get('password', '').strip()

    user = request.user

    try:
        # ===== USER =====
        if username and username != user.username:
            if User.objects.filter(username=username).exclude(id=user.id).exists():
                return JsonResponse({'status': 'error', 'message': 'Bu login (username) band'}, status=400)
            user.username = username

        if email:
            try:
                validate_email(email)
            except ValidationError:
                return JsonResponse({'status': 'error', 'message': 'Email noto‘g‘ri formatda'}, status=400)
            user.email = email

        if admin_name:
            user.first_name = admin_name

        if password:
            if len(password) < 8:
                return JsonResponse({'status': 'error', 'message': 'Parol kamida 8 ta belgidan iborat bo‘lishi kerak'}, status=400)
            user.set_password(password)

        user.save()

        if password:
            update_session_auth_hash(request, user)  # parol o'zgarganda logout bo'lmasligi uchun

        # ===== ADMIN PROFILE =====
        AdminProfile.objects.update_or_create(
            user=user,
            defaults={'full_name': admin_name or user.get_full_name() or user.username}
        )

        return JsonResponse({'status': 'ok', 'message': 'Profil muvaffaqiyatli yangilandi'})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': f'Xatolik: {str(e)}'}, status=500)

def notifications_api(request):
    """Return unseen notifications for admin panel (JSON)."""
    try:
        from suv_tashish_crm.models import Notification
        qs = Notification.objects.filter(seen=False).order_by('-created_at')[:50]
        data = []
        for n in qs:
            data.append({'id': n.id, 'title': n.title, 'message': n.message, 'created_at': n.created_at.isoformat()})
        return JsonResponse({'status': 'ok', 'notifications': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})


def orders_view(request):
    # pop any flash message set by other views (edit/delete)
    flash_message = request.session.pop('flash_message', None)
    flash_type = request.session.pop('flash_type', None)

    # Fetch recent orders with related client/courier/region
    orders_qs = Order.objects.select_related('client__region', 'courier').order_by('-created_at')[:50]

    # include static regions parsed from CSV so templates can show static region info if needed
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []

    # build a mapping from region name -> location (CSV 'location' column)
    static_location_map = {r.get('name'): r.get('location') for r in static_regions}

    orders = []
    for o in orders_qs:
        # ✅ SAFE client fields
        client_obj = getattr(o, "client", None)
        client_name = getattr(client_obj, "full_name", "") if client_obj else ""
        client_phone = getattr(client_obj, "phone", "") if client_obj else ""

        # ✅ SAFE address
        o_address = ""
        try:
            if client_obj and getattr(client_obj, "region", None):
                region_name = getattr(client_obj.region, "name", "")
                o_address = static_location_map.get(region_name) or region_name
        except Exception:
            o_address = ""

        # ✅ SAFE bottles count (Order model har xil bo‘lishi mumkin)
        bottles = (
            getattr(o, "bottle_count", None)
            or getattr(o, "bottles", None)
            or getattr(o, "bottles_count", None)
            or getattr(o, "bottle", None)
            or 0
        )
        try:
            bottles = int(bottles)
        except Exception:
            bottles = 0

        # ✅ status label/class
        if getattr(o, "status", "") == 'done':
            s_label = 'Yetkazildi'
            s_class = 'text-green-600 bg-green-100'
        elif getattr(o, "status", "") in ('delivering', 'assigned'):
            s_label = 'Jarayonda'
            s_class = 'text-yellow-600 bg-yellow-100'
        else:
            s_label = 'Berilmadi'
            s_class = 'text-red-600 bg-red-100'

        # ✅ amounts (payment_amount bo‘lmasa debt_change)
        amount = None
        try:
            if getattr(o, 'payment_amount', None) is not None:
                amount = int(o.payment_amount)
            elif getattr(o, 'debt_change', None) is not None:
                amount = int(o.debt_change)
        except Exception:
            amount = None

        # ✅ payment_type label
        pt = getattr(o, "payment_type", None)
        if pt == "cash":
            payment_type_label = "Pul berdi (naqd)"
        elif pt == "debt":
            payment_type_label = "Qarz bo'ldi"
        elif pt == "click":
            payment_type_label = "Click/online"
        else:
            payment_type_label = ""

        courier_obj = getattr(o, "courier", None)
        courier_name = getattr(courier_obj, "full_name", None) if courier_obj else None

        orders.append({
            'order_id': o.id,
            'order_label': f"#{o.id}",
            'client_id': client_obj.id if client_obj else None,
            'user_id': f"u_{client_obj.id}" if client_obj else None,

            'client_name': client_name,
            'phone': client_phone,
            'address': o_address,

            'lat': getattr(client_obj, 'location_lat', None) if client_obj else None,
            'lon': getattr(client_obj, 'location_lon', None) if client_obj else None,

            'comment': (getattr(o, 'client_note', None) or (getattr(client_obj, "note", "") if client_obj else "")),

            'status': getattr(o, "status", ""),
            'status_label': s_label,
            'status_class': s_class,

            'courier': courier_name,

            # ✅ muhim joy: endi xato bermaydi
            'bottles': bottles,

            'amount': amount,

            'payment_type': payment_type_label,
            'payment_amount': int(getattr(o, 'payment_amount', 0) or 0) if getattr(o, 'payment_amount', None) is not None else None,
        })

    # Top clients by number of orders
    top_clients = (
        Order.objects.values('client__id', 'client__full_name', 'client__phone')
        .annotate(orders_count=Count('id'))
        .order_by('-orders_count')[:10]
    )

    # Top couriers by delivered orders
    top_couriers = (
        Order.objects.filter(status='done')
        .values('courier__id', 'courier__full_name')
        .annotate(delivered=Count('id'))
        .order_by('-delivered')[:10]
    )

    # Orders per region
    region_stats = (
        Order.objects.values('client__region__id', 'client__region__name')
        .annotate(total_orders=Count('id'))
        .order_by('-total_orders')
    )

    # Also provide all client profiles so the Orders page can show "all profiles"
    clients_qs = Client.objects.select_related('region').all()

    profiles = []
    for c in clients_qs:
        address = ''
        if c.region:
            address = static_location_map.get(c.region.name) or getattr(c.region, 'name', '')

        # last order status
        last_order = Order.objects.filter(client=c).order_by('-created_at').first()
        if last_order:
            st = getattr(last_order, "status", "")
            if st == 'done':
                p_label = 'Yetkazildi'
                p_class = 'text-green-600 bg-green-100'
            elif st in ('delivering', 'assigned'):
                p_label = 'Jarayonda'
                p_class = 'text-yellow-600 bg-yellow-100'
            else:
                p_label = 'Berilmadi'
                p_class = 'text-red-600 bg-red-100'
        else:
            p_label = ''
            p_class = 'text-gray-600'

        profiles.append({
            'client_id': c.id,
            'order_label': '',
            'user_id': f"u_{c.id}",
            'client_name': c.full_name,
            'phone': c.phone,
            'address': address,
            'comment': c.note or '',
            'debt': getattr(c, 'debt', Decimal('0.00')) or Decimal('0.00'),
            'status_label': p_label,
            'status_class': p_class,
            'status': '',
            'courier': None,
            'bottles': getattr(c, 'bottle_balance', 0),
        })

    context = {
        'orders': orders,
        'profiles': profiles,
        'static_regions': static_regions,
        'regions': Region.objects.order_by('name'),
        'flash_message': flash_message,
        'flash_type': flash_type,
        'top_clients': top_clients,
        'top_couriers': top_couriers,
        'region_stats': region_stats,
    }
    return render(request, 'admin/orders.html', context)


def seed_region_orders(request):
    """Create one static order per entry in Volidam.csv.

    This view must be called with POST (token-protected). For each CSV row:
    - ensure a Region exists
    - create or get a Client for that region (unique phone generated if missing)
    - create an Order for that client
    """
    if request.session.get('role') != 'admin':
        # require login as admin
        return redirect('/login/')

    if request.method != 'POST':
        return redirect('orders_view')

    static_regions = read_csv_data()
    # Auth guard temporarily disabled; re-enable when ready
    created = 0
    for idx, r in enumerate(static_regions):
        region_name = r.get('name') or f'Region {idx}'
        region_obj, _ = Region.objects.get_or_create(name=region_name)

        # Try to use CSV phone if provided and unique, otherwise create synthetic phone
        phone = r.get('phone') or ''
        if phone:
            # normalize phone length
            phone = phone.strip()
            if len(phone) > 15:
                phone = phone[:15]

        if not phone or Client.objects.filter(phone=phone).exists():
            # generate unique phone-like identifier
            phone = f"r{idx}_{int(timezone.now().timestamp())}"[:15]
            # ensure uniqueness by appending idx if necessary
            while Client.objects.filter(phone=phone).exists():
                phone = f"{phone}_{random.randint(0,999)}"[:15]

        # Use CSV name when available, otherwise a neutral client name
        csv_name = (r.get('name') or '').strip()
        if csv_name:
            client_name = csv_name
        else:
            client_name = f"Mijoz {idx + 1}"

        client, created_client = Client.objects.get_or_create(phone=phone, defaults={
            'full_name': client_name,
            'region': region_obj,
            'location_lat': 0.0,
            'location_lon': 0.0,
        })

        # If client existed but has a generic 'Mijoz N' name and CSV provides a better name, update it
        if not created_client and csv_name and (client.full_name.startswith('Mijoz') or client.full_name.strip() == ''):
            client.full_name = csv_name
            client.save()

        # assign a random debt for some clients (30% chance)
        if created_client:
            if random.random() < 0.3:
                debt_amount = Decimal(str(round(random.uniform(10, 300), 2)))
                client.debt = debt_amount
            else:
                client.debt = Decimal('0.00')
            client.save()

        # create a static order for the client (1-5 bottles randomly)
        bottles = random.randint(1, 5)
        Order.objects.create(
            client=client,
            courier=None,
            bottle_count=bottles,
            client_note=f"Static seed order for {region_name}",
            status='done',
        )

        # update client stats
        client.last_order = timezone.now()
        client.bottle_balance = getattr(client, 'bottle_balance', 0) + 0
        client.save()
        created += 1

    # Redirect back with a very small message via session (optional)
    request.session['seeded_regions'] = created
    return redirect('orders_view')


def edit_client(request, client_id):
    try:
        client = Client.objects.get(pk=client_id)
    except Client.DoesNotExist:
        return redirect('orders_view')

    if request.method == 'POST':
        client.full_name = request.POST.get('full_name', client.full_name)
        phone = request.POST.get('phone', client.phone)
        if phone and phone != client.phone:
            # ensure unique phone - basic check
            if not Client.objects.filter(phone=phone).exclude(pk=client.pk).exists():
                client.phone = phone
        region_id = request.POST.get('region_id', '').strip()
        if region_id:
            try:
                region_obj = Region.objects.get(pk=int(region_id))
                client.region = region_obj
            except (Region.DoesNotExist, ValueError):
                pass
        client.note = request.POST.get('note', client.note)
        client.save()
        # set flash message to show on orders page
        request.session['flash_message'] = 'Mijoz muvaffaqiyatli tahrirlandi.'
        request.session['flash_type'] = 'success'
        return redirect('orders_view')

    # Prepare regions list with CSV 'location' as display text when available
    static_regions = read_csv_data()
    static_map = {r.get('name'): r.get('location') for r in static_regions}
    regions_qs = Region.objects.order_by('name')
    regions_with_location = [(r.id, static_map.get(r.name) or r.name) for r in regions_qs]
    return render(request, 'admin/edit_client.html', {'client': client, 'regions_with_location': regions_with_location})


def delete_client(request, client_id):
    if request.method == 'POST':
        try:
            client = Client.objects.get(pk=client_id)
            client.delete()
        except Client.DoesNotExist:
            pass
    # set flash message for deletion
    request.session['flash_message'] = 'Mijoz o`chirildi.'
    request.session['flash_type'] = 'danger'
    return redirect('orders_view')


def add_courier(request):
    # Add a new courier via simple form
    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        region_id = request.POST.get('region_id', '').strip()
        is_active = True if request.POST.get('is_active') == 'on' else False

        region_obj = None
        if region_id:
            # Accept either an integer PK or a region name from CSV choices
            try:
                region_obj = Region.objects.get(pk=int(region_id))
            except Exception:
                try:
                    region_obj, _ = Region.objects.get_or_create(name=region_id)
                except Exception:
                    region_obj = None

        if full_name and phone:
            Courier.objects.create(full_name=full_name, phone=phone, region=region_obj, is_active=is_active)
            request.session['flash_message'] = 'Kuryer qo\'shildi.'
            request.session['flash_type'] = 'success'
            return redirect('couriers_view')

    # Build region choices: prefer DB Regions, but append unique CSV locations for convenience
    # Build region choices: prefer CSV 'location' values first to avoid CSV-name pollution
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []
    try:
        db_regions = list(Region.objects.order_by('name'))
    except Exception:
        db_regions = []

    seen = set()
    region_choices = []
    # prefer CSV locations (these are actual hudud names in Volidam.csv)
    for r in static_regions:
        loc = (r.get('location') or '').strip()
        if loc and loc not in seen:
            seen.add(loc)
            region_choices.append({'value': loc, 'label': loc})
    # then append DB regions that are not duplicates
    for r in db_regions:
        name = (r.name or '').strip()
        if name and name not in seen:
            seen.add(name)
            region_choices.append({'value': str(r.id), 'label': name})

    return render(request, 'admin/add_edit_courier_clean.html', {'region_choices': region_choices, 'action': 'add'})


def edit_courier(request, courier_id):
    try:
        courier = Courier.objects.get(pk=courier_id)
    except Courier.DoesNotExist:
        request.session['flash_message'] = 'Kuryer topilmadi.'
        request.session['flash_type'] = 'danger'
        return redirect('couriers_view')

    if request.method == 'POST':
        courier.full_name = request.POST.get('full_name', courier.full_name).strip()
        phone = request.POST.get('phone', courier.phone).strip()
        if phone and phone != courier.phone:
            # ensure unique phone among couriers (best effort)
            if not Courier.objects.filter(phone=phone).exclude(pk=courier.pk).exists():
                courier.phone = phone

        region_id = request.POST.get('region_id', '').strip()
        if region_id:
            try:
                courier.region = Region.objects.get(pk=int(region_id))
            except Exception:
                try:
                    region_obj, _ = Region.objects.get_or_create(name=region_id)
                    courier.region = region_obj
                except Exception:
                    pass

        courier.is_active = True if request.POST.get('is_active') == 'on' else False
        courier.save()

        request.session['flash_message'] = 'Kuryer muvaffaqiyatli tahrirlandi.'
        request.session['flash_type'] = 'success'
        return redirect('couriers_view')

    # Build region choices (same as add view)
    try:
        db_regions = list(Region.objects.order_by('name'))
    except Exception:
        db_regions = []
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []

    seen = set()
    region_choices = []
    for r in db_regions:
        name = (r.name or '').strip()
        if name and name not in seen:
            seen.add(name)
            region_choices.append({'value': str(r.id), 'label': name})
    for r in static_regions:
        loc = (r.get('location') or '').strip()
        if loc and loc not in seen:
            seen.add(loc)
            region_choices.append({'value': loc, 'label': loc})

    return render(request, 'admin/add_edit_courier_clean.html', {'region_choices': region_choices, 'courier': courier, 'action': 'edit'})

from django.shortcuts import redirect, get_object_or_404

def delete_courier(request, courier_id):
    if request.method == "POST":
        try:
            courier = Courier.objects.get(pk=courier_id)
            courier.delete()
            request.session["flash_message"] = "Kuryer o‘chirildi."
            request.session["flash_type"] = "danger"
        except Courier.DoesNotExist:
            request.session["flash_message"] = "Kuryer topilmadi."
            request.session["flash_type"] = "danger"

    # 🔥 MUHIM JOY — namespace bilan redirect
    return redirect("admin_panel:couriers_view")


def reports_view(request):
    # Build weekly and monthly revenue datasets and top couriers for the template
    import json
    from django.db.models import Q

    try:
        today = timezone.localdate()
    except Exception:
        today = datetime.date.today()

    # Weekly: last 7 days (labels and sums)
    weekly_labels = []
    weekly_data = []
    try:
        for i in range(6, -1, -1):
            day = today - datetime.timedelta(days=i)
            weekly_labels.append(day.strftime('%a'))
            # Sum payment_amount for orders delivered that day (prefer delivered_at, fallback to created_at)
            qs = Order.objects.filter(
                Q(delivered_at__date=day) | (Q(status='done') & Q(created_at__date=day))
            )
            total = qs.aggregate(total=Sum('payment_amount'))
            val = total.get('total')
            # fallback to count of delivered orders if payment_amount not set
            if not val:
                val = qs.count()
            try:
                weekly_data.append(int(val))
            except Exception:
                weekly_data.append(0)
    except Exception:
        weekly_labels = []
        weekly_data = []

    # Monthly: last 12 months
    monthly_labels = []
    monthly_data = []
    try:
        for i in range(11, -1, -1):
            ref = (today.replace(day=1) - datetime.timedelta(days=1)) - datetime.timedelta(days=30 * i)
            year = ref.year
            month = ref.month
            monthly_labels.append(ref.strftime('%b'))
            qs = Order.objects.filter(
                Q(delivered_at__year=year, delivered_at__month=month) | (Q(status='done') & Q(created_at__year=year, created_at__month=month))
            )
            total = qs.aggregate(total=Sum('payment_amount'))
            val = total.get('total')
            if not val:
                val = qs.count()
            try:
                monthly_data.append(int(val))
            except Exception:
                monthly_data.append(0)
    except Exception:
        monthly_labels = []
        monthly_data = []

    # Top couriers: percentage score relative to top performer (use delivered counts)
    top_couriers = []
    try:
        qs = (
            Order.objects.filter(status='done')
            .values('courier__id', 'courier__full_name')
            .annotate(delivered=Count('id'))
            .order_by('-delivered')[:5]
        )
        max_delivered = 0
        for i, t in enumerate(qs):
            if i == 0:
                max_delivered = t.get('delivered', 0) or 0
            top_couriers.append({'id': t.get('courier__id'), 'name': t.get('courier__full_name') or '—', 'delivered': t.get('delivered', 0)})
        for c in top_couriers:
            try:
                c['score'] = int((c['delivered'] / max_delivered) * 100) if max_delivered else 0
            except Exception:
                c['score'] = 0
    except Exception:
        top_couriers = []

    context = {
        'weekly_labels_json': json.dumps(weekly_labels, ensure_ascii=False),
        'weekly_data_json': json.dumps(weekly_data),
        'monthly_labels_json': json.dumps(monthly_labels, ensure_ascii=False),
        'monthly_data_json': json.dumps(monthly_data),
        'top_couriers': top_couriers,
    }
    return render(request, 'admin/reports.html', context)


def courier_ranking_view(request):
    """Show courier ranking based on number of delivered orders."""
    try:
        qs = (
            Order.objects.filter(status='done')
            .values('courier__id', 'courier__full_name')
            .annotate(delivered=Count('id'))
            .order_by('-delivered')
        )
        ranking = []
        max_delivered = 0
        for i, t in enumerate(qs):
            if i == 0:
                max_delivered = t.get('delivered', 0) or 0
            ranking.append({
                'id': t.get('courier__id'),
                'name': t.get('courier__full_name') or '—',
                'delivered': t.get('delivered', 0),
            })
    except Exception:
        ranking = []
        max_delivered = 0

    # compute score percent relative to top performer
    for r in ranking:
        try:
            r['score_percent'] = int((r['delivered'] / max_delivered) * 100) if max_delivered else 0
        except Exception:
            r['score_percent'] = 0

    return render(request, 'admin/courier_ranking.html', {'ranking': ranking})


def debtors_view(request):
    """Show clients with outstanding debt as profile cards.

    This view lists `Client` records with `debt > 0`. For each client we include:
    - last order id (if any)
    - user id label
    - phone
    - courier who served last order (if any)
    - debt amount
    - status (derived from last order)
    - days overdue (based on client's last_order timestamp if available)
    """
    from django.utils import timezone

    title = 'Qarzdorlar'
    # Optional filters from query params
    filter_name = (request.GET.get('name') or '').strip()
    filter_phone = (request.GET.get('phone') or '').strip()
    filter_region = (request.GET.get('region') or '').strip()

    # Base queryset: all clients with non-zero debt (positive or negative)
    from django.db.models import Q as _Q
    clients_qs = Client.objects.exclude(debt=0).select_related('region')
    debtors = []
    # Apply filters if provided (robust fallbacks)
    try:
        if filter_name:
            clients_qs = clients_qs.filter(
                Q(full_name__icontains=filter_name) | Q(first_name__icontains=filter_name) | Q(last_name__icontains=filter_name)
            )
        if filter_phone:
            digits = ''.join(ch for ch in filter_phone if ch.isdigit())
            if digits:
                clients_qs = clients_qs.filter(Q(phone__icontains=filter_phone) | Q(phone__icontains=digits))
            else:
                clients_qs = clients_qs.filter(phone__icontains=filter_phone)
        if filter_region:
            try:
                rid = int(filter_region)
                clients_qs = clients_qs.filter(region__id=rid)
            except Exception:
                clients_qs = clients_qs.filter(Q(region__name__icontains=filter_region))
    except Exception:
        clients_qs = Client.objects.filter(debt__gt=0).select_related('region')
    now = timezone.now()
    # load static CSV mapping (phone -> csv name) to prefer CSV name when available
    try:
        static_regions = read_csv_data()
    except Exception:
        static_regions = []

    # build mapping from CSV region name -> CSV location (address)
    static_location_map = { r.get('name'): r.get('location') for r in static_regions }

    def _normalize_to_998(raw_digits: str):
        """Normalize any digits string to the form '998' + 9 local digits (12 digits total).
        Returns the normalized digits (without '+') or None if not possible.
        Strategy: take the last 9 digits as local number and prepend '998'.
        """
        if not raw_digits:
            return None
        d = ''.join(ch for ch in raw_digits if ch.isdigit())
        if len(d) < 9:
            return None
        last9 = d[-9:]
        std = '998' + last9
        if len(std) == 12:
            return std
        return None

    csv_phone_map = {}
    for r in static_regions:
        p = (r.get('phone') or '').strip()
        norm = _normalize_to_998(p)
        if norm:
            csv_phone_map[norm] = r.get('name')

    # build a fallback mapping by last-6-digits to catch phones with different formatting
    csv_phone_by_last6 = {}
    for r in static_regions:
        p = ''.join(ch for ch in (r.get('phone') or '') if ch.isdigit())
        if len(p) >= 6:
            csv_phone_by_last6[p[-6:]] = r.get('name')

    # Prefer real couriers from DB; fallback to a small static list if none
    try:
        couriers_qs = list(Courier.objects.all())
        if couriers_qs:
            couriers_list = [{'name': c.full_name, 'phone': getattr(c, 'phone', '-') } for c in couriers_qs]
        else:
            couriers_list = [
                {'name': 'Oybek Kurye', 'phone': '+998901112233'},
                {'name': 'Aziz Kurye', 'phone': '+998902223344'},
                {'name': 'Dilorom Kurye', 'phone': '+998903334455'},
            ]
    except Exception:
        couriers_list = [
            {'name': 'Oybek Kurye', 'phone': '+998901112233'},
            {'name': 'Aziz Kurye', 'phone': '+998902223344'},
            {'name': 'Dilorom Kurye', 'phone': '+998903334455'},
        ]

    def _is_static_client_phone(raw_phone: str) -> bool:
        """Return True if phone looks like a synthetic/static phone created by seeding.

        Heuristics:
        - phone strings generated by seeding start with 'r' or contain underscores
        - phone must contain at least 9 digits to be considered valid (local number)
        Note: If phone is empty/null we consider it *not* static so admins can see
        debtors without phone values.
        """
        if not raw_phone:
            return False
        try:
            if isinstance(raw_phone, str) and raw_phone.startswith('r'):
                return True
        except Exception:
            pass
        if '_' in str(raw_phone):
            return True
        digits = ''.join(ch for ch in (raw_phone or '') if ch.isdigit())
        if len(digits) < 9:
            return True
        return False

    for c in clients_qs:
        # skip synthetic/static seeded clients so admin shows only real clients
        try:
            if _is_static_client_phone(c.phone):
                continue
        except Exception:
            # on unexpected error, skip this client to avoid showing bad data
            continue
        last_order = Order.objects.filter(client=c).order_by('-created_at').first()
        order_id = last_order.id if last_order else None
        # Determine courier info: prefer order.courier if present, otherwise pick a static courier
        if last_order and last_order.courier:
            courier_name = last_order.courier.full_name
            courier_phone = getattr(last_order.courier, 'phone', '-')
        else:
            # pick a courier deterministically by client id from the DB-derived list
            sc = couriers_list[c.id % len(couriers_list)]
            courier_name = sc['name']
            courier_phone = sc['phone']
        # determine status label
        if last_order:
            st = last_order.status
            if st == 'done':
                s_label = 'Yetkazildi'
            elif st in ('delivering', 'assigned'):
                s_label = 'Jarayonda'
            else:
                s_label = 'Berilmadi'
        else:
            s_label = ''

        # days overdue: compare now and client's last_order
        days_overdue = 0
        if getattr(c, 'last_order', None):
            try:
                delta = now.date() - c.last_order.date()
                days_overdue = delta.days
            except Exception:
                days_overdue = 0

        # Format phone: normalize to +998XXXXXXXXX (13 chars including '+') when possible
        raw_phone = (c.phone or '')
        norm = _normalize_to_998(raw_phone)
        if norm:
            formatted_phone = '+' + norm
        else:
            # fallback: strip non-digits and show what's available
            digits = ''.join(ch for ch in raw_phone if ch.isdigit())
            formatted_phone = digits or raw_phone

        # If CSV has a name for this phone, prefer it. Also try last-6-digits fallback.
        csv_name = None
        if norm and norm in csv_phone_map:
            csv_name = csv_phone_map.get(norm)
        else:
            # try fallback by last-6-digits
            digits_only = ''.join(ch for ch in raw_phone if ch.isdigit())
            if len(digits_only) >= 6:
                last6 = digits_only[-6:]
                if last6 in csv_phone_by_last6:
                    csv_name = csv_phone_by_last6.get(last6)

        # If we found a better name in CSV and client currently has a generic 'Mijoz N' name, update DB
        try:
            if csv_name and (c.full_name.startswith('Mijoz') or c.full_name.strip() == ''):
                c.full_name = csv_name
                c.save()
        except Exception:
            # If save fails for any reason, continue without interrupting the listing
            pass

        # determine address: prefer CSV location for client's region name
        address = ''
        if c.region:
            address = static_location_map.get(c.region.name) or getattr(c.region, 'name', '')

        # last order date (when they took the last order)
        last_order_date = ''
        if last_order and getattr(last_order, 'created_at', None):
            try:
                last_order_date = last_order.created_at.strftime('%Y-%m-%d %H:%M')
            except Exception:
                last_order_date = str(last_order.created_at)

        debtors.append({
            'client_id': c.id,
            'order_id': order_id,
            'user_id': f"u_{c.id}",
            'name': csv_name or c.full_name,
            'phone': formatted_phone,
                'courier': courier_name,
                'courier_phone': courier_phone,
            'debt': getattr(c, 'debt', 0),
            'status_label': s_label,
            'days_overdue': days_overdue,
            'address': address,
            'last_order_date': last_order_date,
        })

    # Pop any flash message set by previous actions (e.g., mark_debtor_paid)
    flash_message = request.session.pop('flash_message', None)
    flash_type = request.session.pop('flash_type', None)

    # Regions list for the filter select (include DB regions)
    try:
        regions_list = list(Region.objects.order_by('name'))
    except Exception:
        regions_list = []

    return render(request, 'admin/debtors.html', {
        'debtors': debtors,
        'title': title,
        'flash_message': flash_message,
        'flash_type': flash_type,
        'regions': regions_list,
        'filter_name': filter_name,
        'filter_phone': filter_phone,
        'filter_region': filter_region,
    })


def mark_debtor_paid(request, client_id):
    if request.method == 'POST':
        try:
            client = Client.objects.get(pk=client_id)
            client.debt = 0
            client.save()
            request.session['flash_message'] = f"{client.full_name} - qarz to'landi."
            request.session['flash_type'] = 'success'
        except Client.DoesNotExist:
            request.session['flash_message'] = 'Mijoz topilmadi.'
            request.session['flash_type'] = 'danger'
    return redirect('admin_panel:debtors_view')


def inactive_clients_view(request):
    """List clients who have not had an order for `cutoff_days` or more.

    Default cutoff is 30 days, but can be overridden with ?cutoff=<days>.
    """
    from django.utils import timezone

    try:
        cutoff_days = int(request.GET.get('cutoff', 30))
    except Exception:
        cutoff_days = 30

    cutoff_date = timezone.now() - datetime.timedelta(days=cutoff_days)

    # Select clients whose last_order is null (never ordered) or older than cutoff
    try:
        qs = Client.objects.filter(Q(last_order__isnull=True) | Q(last_order__lte=cutoff_date)).order_by('last_order')[:100]
    except Exception:
        qs = Client.objects.none()

    clients = []
    now = timezone.now()
    for c in qs:
        if getattr(c, 'last_order', None):
            try:
                last_date = c.last_order.strftime('%Y-%m-%d')
                days_since = (now.date() - c.last_order.date()).days
            except Exception:
                last_date = str(c.last_order)
                days_since = None
        else:
            last_date = None
            days_since = None

        phone = c.phone or ''
        clients.append({
            'client_id': c.id,
            'name': c.full_name,
            'phone': phone,
            'last_order_date': last_date,
            'days_since': days_since,
        })

    return render(request, 'admin/inactive_clients.html', {'clients': clients, 'cutoff_days': cutoff_days})


@csrf_exempt
def delete_order(request, order_id):
    """Delete an order by ID."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)
    
    try:
        order = Order.objects.get(pk=order_id)
        order.delete()
        return JsonResponse({'status': 'ok', 'message': 'Buyurtma o\'chirildi'})
    except Order.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Buyurtma topilmadi'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

from django.contrib import messages
from common.csv_utils import append_client_to_csv


def add_client(request):
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        phone = request.POST.get("phone", "").strip()
        address = request.POST.get("address", "").strip()

        # minimal validatsiya
        if not full_name or not phone or not address:
            messages.error(request, "Iltimos, barcha maydonlarni to‘ldiring.")
            return render(request, "add_client.html")

        csv_path = append_client_to_csv(full_name, phone, address, source="admin")


        messages.success(request, f"✅ Mijoz saqlandi. CSV: {csv_path}")

        return redirect("admin_panel:admin_dashboard")

    return render(request, "admin_panel/add_client.html")
import os
import csv
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def admin_clients_export_excel(request):
    """
    Volidam.csv fayldan ma'lumotlarni olib .xlsx qilib export qiladi
    """
    # ✅ CSV fayl yo'li (Volidam.csv project rootda bo'lsa)
    csv_path = os.path.join(settings.BASE_DIR, "Volidam.csv")

    if not os.path.exists(csv_path):
        return HttpResponse(f"CSV topilmadi: {csv_path}", status=404)

    # ✅ CSV o'qish (utf-8-sig: Excel/Windowsdan chiqqan CSVlarda yaxshi ishlaydi)
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # ✅ Excel yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Headerlar (CSVdagi headerlar qanday bo'lsa, shuni olamiz)
    headers = reader.fieldnames or []
    if not headers:
        return HttpResponse("CSV header topilmadi (1-qator bo'sh bo'lishi mumkin).", status=400)

    # Header yozish
    ws.append(headers)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data yozish
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    # ✅ Column width auto
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)

    # ✅ Response (download)
    filename = f"clients_export_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib import messages

from suv_tashish_crm.models import Client, Region

# senlarda bor funksiya (CSV o‘qiydi) - shuni ishlatamiz
# from .utils import read_csv_data  # agar boshqa faylda bo‘lsa
# yoki shu faylda oldin mavjud bo‘lsa, import qilma

def clients_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = Client.objects.select_related("region").order_by("-id")
    if q:
        qs = qs.filter(full_name__icontains=q) | qs.filter(phone__icontains=q)

    return render(request, "admin/clients.html", {
        "clients": qs[:500],
        "q": q,
    })


@require_http_methods(["POST"])
def clients_import_csv(request):
    """
    CSV'dagi mijozlarni DB'ga bir marta import qiladi.
    Telefon bo‘yicha duplicate bo‘lsa update qilmaydi (xohlasang update ham qilamiz).
    """
    try:
        rows = read_csv_data()  # ✅ sening mavjud funksiyang
    except Exception as e:
        messages.error(request, f"CSV o‘qishda xato: {e}")
        return redirect("admin_panel:clients_list")

    created = 0
    skipped = 0

    for r in rows:
        name = (r.get("full_name") or r.get("name") or "").strip()
        phone = (r.get("phone") or "").strip()
        region_name = (r.get("region") or r.get("region_name") or "").strip()

        if not phone:
            skipped += 1
            continue

        region_obj = None
        if region_name:
            region_obj = Region.objects.filter(name__iexact=region_name).first()

        obj, was_created = Client.objects.get_or_create(
            phone=phone,
            defaults={
                "full_name": name,
                "region": region_obj,
            }
        )

        if was_created:
            created += 1
        else:
            # bor bo‘lsa ham bo‘sh fields ni to‘ldirib qo‘yamiz (ixtiyoriy)
            updated = False
            if name and not obj.full_name:
                obj.full_name = name
                updated = True
            if region_obj and not obj.region:
                obj.region = region_obj
                updated = True
            if updated:
                obj.save()
            skipped += 1

    messages.success(request, f"CSV import: {created} ta qo‘shildi, {skipped} ta o‘tkazildi.")
    return redirect("admin_panel:clients_list")

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from suv_tashish_crm.models import Client, Region
import os
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
import os

def read_couriers_xlsx_locations():
    # ✅ couriers.xlsx joyi:
    xlsx_path = os.path.join(settings.BASE_DIR, "couriers.xlsx")
    # agar "data" papkada bo‘lsa buni ishlat:
    # xlsx_path = os.path.join(settings.BASE_DIR, "data", "couriers.xlsx")

    print("XLSX PATH:", xlsx_path)  # ✅ debug
    if not os.path.exists(xlsx_path):
        print("❌ couriers.xlsx topilmadi!")
        return []

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # headerlar
    headers = []
    for cell in ws[1]:
        headers.append((str(cell.value).strip().lower() if cell.value else ""))

    print("HEADERS:", headers)  # ✅ debug

    candidates = ["hudud", "manzil", "location", "region", "address"]
    col_idx = None
    for i, h in enumerate(headers):
        if any(k in h for k in candidates):
            col_idx = i + 1
            break

    # agar header topilmasa -> 1-ustun deb olamiz
    if col_idx is None:
        col_idx = 1

    seen = set()
    locations = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        loc = (str(val).strip() if val is not None else "")
        if loc and loc not in seen:
            seen.add(loc)
            locations.append(loc)

    print("LOC COUNT:", len(locations))  # ✅ debug
    return locations

def edit_client(request, client_id):
    client = get_object_or_404(Client.objects.select_related("region"), pk=client_id)

    if request.method == "POST":
        full_name = (request.POST.get("full_name") or "").strip()
        phone = (request.POST.get("phone") or "").strip()
        region_id = (request.POST.get("region_id") or request.POST.get("region") or "").strip()
        note = (request.POST.get("note") or "").strip()

        if phone and Client.objects.exclude(pk=client.pk).filter(phone=phone).exists():
            messages.error(request, "Bu telefon raqam boshqa mijozda bor.")
            return redirect("admin_panel:edit_client", client_id=client.pk)

        client.full_name = full_name
        client.phone = phone

        # ✅ Hudud saqlash (DB id yoki couriers.xlsx)
        if region_id.startswith("xlsx:"):
            loc_name = region_id.replace("xlsx:", "", 1).strip()
            if loc_name:
                region_obj, _ = Region.objects.get_or_create(name=loc_name)
                client.region = region_obj
            else:
                client.region = None
        elif region_id:
            # db id bo‘lsa
            try:
                client.region = Region.objects.filter(id=int(region_id)).first()
            except Exception:
                client.region = None
        else:
            client.region = None

        if hasattr(client, "note"):
            client.note = note

        client.save()
        messages.success(request, "Mijoz yangilandi.")
        return redirect("admin_panel:clients_list")

    # ✅ GET: select uchun list tayyorlaymiz
    db_regions = list(Region.objects.order_by("name"))
    xlsx_locations = []
    try:
        xlsx_locations = read_couriers_xlsx_locations()
    except Exception:
        xlsx_locations = []

    regions_with_location = []
    # 1) DB
    for r in db_regions:
        regions_with_location.append((str(r.id), r.name))
    # 2) XLSX
    for loc in xlsx_locations:
        regions_with_location.append((f"xlsx:{loc}", loc))

    return render(request, "admin/edit_client.html", {
        "client": client,
        "regions_with_location": regions_with_location,
    })


@require_http_methods(["POST"])
def delete_client(request, client_id):
    c = get_object_or_404(Client, pk=client_id)
    c.delete()
    messages.success(request, "Mijoz o‘chirildi.")
    return redirect("admin_panel:clients_list") 
import re
from openpyxl import load_workbook
from django.shortcuts import redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from suv_tashish_crm.models import Client
def _norm_phone(phone) -> str:
    if not phone or str(phone).strip() in ["None", ""]: return ""
    s = str(phone).strip().replace(" ", "").replace("-", "").replace("(", "").replace(")", "").replace(".0", "")
    digits = re.sub(r"\D", "", s)
    if digits.startswith("998") and len(digits) >= 12: return "+" + digits[:12]
    if len(digits) == 9: return "+998" + digits
    if s.startswith("+") and len(digits) >= 12: return "+" + digits[:12]
    return ""

def _is_admin(request) -> bool:
    return bool(request.session.get("admin_id"))

@require_POST
def admin_clients_upload_excel(request):
    if not _is_admin(request):
        return redirect("admin_panel:clients_list")

    # Joriy admin biznesini aniqlash
    my_business = _get_my_business(request.user)
    if not my_business:
        messages.error(request, "Sizga biznes biriktirilmagan! Admin panelda biznesni tekshiring.")
        return redirect("admin_panel:clients_list")

    f = request.FILES.get("excel_file")
    if not f or not f.name.lower().endswith(".xlsx"):
        messages.error(request, "Faqat .xlsx formatdagi faylni tanlang.")
        return redirect("admin_panel:clients_list")

    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=f, data_only=True)
        ws = wb.active
        
        # Headerlarni tekshirish
        expected = ["full_name", "bottle_soni", "manzili", "phone"]
        header = [str(ws.cell(row=1, column=i).value).strip().lower() for i in range(1, 5)]

        if header != expected:
            messages.error(request, f"Excel sarlavhasi xato! Kerakli: {', '.join(expected)}")
            return redirect("admin_panel:clients_list")

        created, skipped = 0, 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row): continue
            
            name = str(row[0]).strip() if row[0] else ""
            phone_raw = row[3]
            phone_norm = _norm_phone(phone_raw) # Telefonni formatlash

            if not name:
                skipped += 1
                continue

            # Faqat shu biznes ichida takrorlanishni tekshirish
            if phone_norm and Client.objects.filter(phone=phone_norm, business=my_business).exists():
                skipped += 1
                continue

            try:
                # Bottle sonini o'girish
                try:
                    b_val = int(float(row[1])) if row[1] not in (None, "") else 1
                except:
                    b_val = 1

                # Client modeliga moslab saqlash
                Client.objects.create(
                    business=my_business,  # 👈 Multi-tenancy ulanishi
                    full_name=name,
                    phone=phone_norm if phone_norm else f"no_phone_{uuid.uuid4().hex[:6]}",
                    note=str(row[2]).strip() if row[2] else "", # 'manzili' -> 'note'ga tushadi
                    bottle_balance=b_val,
                    must_change_password=True
                )
                created += 1
            except Exception as e:
                print(f"Row error: {e}") # Debug uchun
                skipped += 1

        messages.success(request, f"Muvaffaqiyatli: {created} ta mijoz. O'tkazildi: {skipped} ta.")
    except Exception as e:
        messages.error(request, f"Tizim xatosi: {str(e)}")
    
    return redirect("admin_panel:clients_list")